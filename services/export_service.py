# -*- coding: utf-8 -*-
"""
导出服务（增强版，支持“多用户一起导出”为 Excel）：

- 保留并增强原有接口：
  - export_user_records(user_id_or_username, ...)   单用户流水
  - export_all_records(...)                          全量流水
  - export_all_users_detail(fmt="xlsx")              所有用户“详细资料”总表
  - export_one_user_full(user_id_or_username, ...)   单用户专用 Excel（Summary + Ledger）
  - export_all_users_and_ledger(fmt="xlsx")          一个 Excel（Users 全量 + Ledger 全量）

- 新增（为 admin "multi export" 提供的统一入口）：
  - export_some_users_and_ledger(*args, **kwargs)
      解析传入的 tg_ids / user_ids / ids / members 等参数（字符串或列表均可），
      仅导出所选用户（Users + Ledger 两个工作表），返回生成的 Excel 路径。

统一增强：
- Users 工作表新增列：nickname（用户昵称）、last_seen_at（最后上线时间）
- Ledger 工作表新增列：nickname（用户昵称）；在合并导出中同时新增 last_seen_at
- 表头：第 1 行中文硬编码（xlsx），第 2 行英文列名；冻结表头 + 启用筛选
- 自动设置列宽（取前若干行估算）

依赖：pandas、openpyxl、SQLAlchemy。模型路径基于你当前项目结构：
  models.db.get_session / models.user.User / models.ledger.Ledger, LedgerType
如你的路径不同，请据实调整本文件顶部的三处导入。
"""

from __future__ import annotations
import os
import re
import enum
from typing import Optional, Iterable, List, Dict, Any, Union, Tuple, Generator, Sequence
from datetime import datetime, timedelta
from decimal import Decimal

from sqlalchemy import func

# 你的项目模型（如路径不同，请在此处改为你项目的实际模块路径）
from models.db import get_session
from models.user import User
from models.ledger import Ledger, LedgerType

# 第三方依赖
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ---------------------------------------
# 通用配置/工具
# ---------------------------------------

EXPORT_DIR = os.path.join(os.getcwd(), "exports")


def _ensure_export_dir() -> str:
    if not os.path.isdir(EXPORT_DIR):
        os.makedirs(EXPORT_DIR, exist_ok=True)
    return EXPORT_DIR


def _dtfmt(dt: Optional[datetime]) -> str:
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _enum_to_str(v: Union[enum.Enum, str, int, None]) -> str:
    if v is None:
        return ""
    if isinstance(v, enum.Enum):
        return v.name
    return str(v)


def _as_float(x: Any) -> float:
    try:
        if isinstance(x, Decimal):
            return float(x)
        return float(x)
    except Exception:
        try:
            return float(str(x))
        except Exception:
            return 0.0


def _normalize_username(u: Optional[str]) -> str:
    if not u:
        return ""
    s = str(u)
    return s[1:] if s.startswith("@") else s


def _display_full_name(
    username: Optional[str] = "",
    full_name: Optional[str] = "",
    first_name: Optional[str] = "",
    last_name: Optional[str] = "",
    name_fallback: Optional[str] = "",
) -> str:
    fn = (full_name or "").strip()
    if fn:
        return fn
    fp = f"{first_name or ''} {last_name or ''}".strip()
    if fp:
        return fp
    return (name_fallback or username or "").strip()


def _as_int_list(x: Any) -> List[int]:
    """把各种形式的参数解析成 int list。
    支持：[1,2,3] / (1,2) / "1, 2, 3" / {"1","2"} / 单个数字字符串 / 命令行 argv 列表。
    """
    if x is None:
        return []
    if isinstance(x, (list, tuple, set)):
        out: List[int] = []
        for v in x:
            try:
                out.append(int(str(v).strip()))
            except Exception:
                pass
        return out
    s = str(x)
    # 如果是从命令行传入的 argv 列表，可能是 "['1','2']" 这种；尽量宽松解析
    if s.startswith("[") and s.endswith("]"):
        s = s.strip("[]")
    parts = re.split(r"[\s,;]+", s.strip())
    out: List[int] = []
    for p in parts:
        if not p:
            continue
        try:
            out.append(int(p))
        except Exception:
            pass
    return out


# ---------- 取 ledger ----------

def _pick_user_by_query(q: Union[str, int]) -> Optional[User]:
    """
    q 可以是：
    - 纯数字或 int：按 Telegram user_id（tg_id）查
    - 以 @ 开头：按 username 查
    - 普通字符串：按 username 或 name 查
    """
    if isinstance(q, int):
        q = str(q)
    q = (q or "").strip()
    if not q:
        return None

    with get_session() as s:
        if q.isdigit():
            u = s.query(User).filter(User.tg_id == int(q)).first()
            if u:
                return u
        name = _normalize_username(q)
        try:
            u = s.query(User).filter(User.username == name).first()
            if u:
                return u
        except Exception:
            pass
        try:
            u = s.query(User).filter(User.name == name).first()
            if u:
                return u
        except Exception:
            pass
        return None


def _query_ledgers(user: Optional[User],
                   start: Optional[datetime],
                   end: Optional[datetime],
                   tokens: Optional[Iterable[str]],
                   types: Optional[Iterable[Union[LedgerType, str]]]) -> List[Ledger]:
    with get_session() as s:
        q = s.query(Ledger)
        if user is not None:
            try:
                q = q.filter(Ledger.user_tg_id == user.tg_id)
            except Exception:
                q = q.filter(Ledger.user_id == getattr(user, "id", None))

        if start is not None:
            q = q.filter(Ledger.created_at >= start)
        if end is not None:
            q = q.filter(Ledger.created_at < end)

        if tokens:
            toks = [str(t).upper() for t in tokens]
            q = q.filter(Ledger.token.in_(toks))

        if types:
            norm_types: List[str] = []
            for t in types:
                if isinstance(t, LedgerType):
                    norm_types.append(t.name)
                else:
                    norm_types.append(str(t).upper())
            try:
                q = q.filter(Ledger.type.in_(norm_types))
            except Exception:
                back: List[LedgerType] = []
                for n in norm_types:
                    try:
                        back.append(LedgerType[n])
                    except Exception:
                        pass
                if back:
                    q = q.filter(Ledger.type.in_(back))

        q = q.order_by(Ledger.created_at.asc(), Ledger.id.asc())
        return q.all()


# ---------- ledger 行 → dict ----------

def _rows_to_records(rows: List[Ledger]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for r in rows:
        d: Dict[str, Any] = {
            "id": getattr(r, "id", None),
            "user_tg_id": getattr(r, "user_tg_id", None) if hasattr(r, "user_tg_id") else getattr(r, "user_id", None),
            "token": (getattr(r, "token", "") or "").upper(),
            "amount": _as_float(getattr(r, "amount", 0)),
            "type": _enum_to_str(getattr(r, "type", "")),
            "note": getattr(r, "note", "") or "",
            "created_at": _dtfmt(getattr(r, "created_at", None)),
        }
        out.append(d)
    return out


# ---------- 批量补齐用户字段到记录（ledger） ----------

def _attach_user_columns(records: List[Dict[str, Any]]) -> None:
    """
    通过 user_tg_id 批量追加用户信息：
      - username / name
      - first_name / last_name / full_name
      - nickname（新增，优先 full_name）
    """
    tg_ids = {r.get("user_tg_id") for r in records if r.get("user_tg_id")}
    if not tg_ids:
        return
    with get_session() as s:
        users: List[User] = s.query(User).filter(User.tg_id.in_(list(tg_ids))).all()
        u_map: Dict[int, Dict[str, Any]] = {}
        for u in users:
            tg_id = int(getattr(u, "tg_id"))
            username = getattr(u, "username", None) or getattr(u, "name", "") or ""
            first_name = getattr(u, "first_name", "") or ""
            last_name  = getattr(u, "last_name", "") or ""
            full_name  = getattr(u, "full_name", "") or ""
            if not full_name:
                fn = f"{first_name} {last_name}".strip()
                full_name = fn if fn else (getattr(u, "name", "") or username or "")
            nickname = _display_full_name(username=username, full_name=full_name, first_name=first_name, last_name=last_name, name_fallback=getattr(u, "name", ""))

            u_map[tg_id] = {
                "username": username or "",
                "first_name": first_name,
                "last_name": last_name,
                "full_name": full_name,
                "nickname": nickname,
            }
    for r in records:
        tg_id = r.get("user_tg_id")
        if tg_id and tg_id in u_map:
            r.update(u_map[tg_id])


# ---------- Excel/CSV 写入与列宽 ----------

def _autosize_columns(ws, start_row: int = 1, max_scan_rows: int = 1000, min_width: int = 6, max_width: int = 60) -> None:
    """根据内容估算列宽（扫描前 max_scan_rows 行）。"""
    dims: Dict[int, int] = {}
    scanned = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        scanned += 1
        if scanned > max_scan_rows:
            break
        for idx, value in enumerate(row, 1):
            length = len(str(value)) if value is not None else 0
            if length > dims.get(idx, 0):
                dims[idx] = length
    for idx, length in dims.items():
        # 中文可能更宽，给一定余量
        width = min(max(min_width, length + 2), max_width)
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_dataframe(
    df: pd.DataFrame,
    basename: str,
    fmt: str = "xlsx",
    header_zh: Optional[List[str]] = None,
    sheet_name: str = "records",
    freeze_header: bool = True,
    enable_filter: bool = True,
) -> str:
    _ensure_export_dir()
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    if fmt.lower() == "csv":
        path = os.path.join(EXPORT_DIR, f"{basename}_{ts}.csv")
        df.to_csv(path, index=False, encoding="utf-8-sig")
        return path

    path = os.path.join(EXPORT_DIR, f"{basename}_{ts}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    cur_row = 1
    if header_zh:
        ws.append(header_zh)
        cur_row += 1

    ws.append(list(df.columns))
    header_row_eng = cur_row
    cur_row += 1

    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
        cur_row += 1

    if enable_filter:
        last_col_letter = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A{header_row_eng}:{last_col_letter}{header_row_eng}"
    if freeze_header:
        ws.freeze_panes = f"A{3 if header_zh else 2}"

    _autosize_columns(ws, start_row=1)

    wb.save(path)
    return path


# ---------------------------------------
# 原有接口（保留、增强）
# ---------------------------------------

def export_user_records(
    user_id_or_username: str,
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    tokens: Optional[Iterable[str]] = None,
    types: Optional[Iterable[Union[LedgerType, str]]] = None,
    fmt: str = "xlsx",
) -> Optional[str]:
    user = _pick_user_by_query(user_id_or_username)
    if not user:
        df = pd.DataFrame([{"error": f"user not found: {user_id_or_username}"}])
        return _write_dataframe(df, "export_user_not_found", fmt)

    rows = _query_ledgers(user, start, end, tokens, types)
    if not rows:
        return None

    recs = _rows_to_records(rows)
    _attach_user_columns(recs)

    df = pd.DataFrame(recs, columns=[
        "id", "created_at", "user_tg_id",
        "username", "first_name", "last_name", "full_name", "nickname",
        "type", "token", "amount", "note"
    ])
    basename = f"export_user_{getattr(user, 'tg_id', 'unknown')}"
    return _write_dataframe(df, basename, fmt)


def export_all_records(
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    tokens: Optional[Iterable[str]] = None,
    types: Optional[Iterable[Union[LedgerType, str]]] = None,
    fmt: str = "xlsx",
) -> Optional[str]:
    rows = _query_ledgers(user=None, start=start, end=end, tokens=tokens, types=types)
    if not rows:
        return None

    recs = _rows_to_records(rows)
    _attach_user_columns(recs)

    df = pd.DataFrame(recs, columns=[
        "id", "created_at", "user_tg_id",
        "username", "first_name", "last_name", "full_name", "nickname",
        "type", "token", "amount", "note"
    ])
    basename = "export_all"
    return _write_dataframe(df, basename, fmt)


# ---------------------------------------
# 用户资料相关
# ---------------------------------------

def _get_user_field(u: User, *candidates: str, default=None):
    for name in candidates:
        if hasattr(u, name):
            return getattr(u, name)
    return default


def _user_balances(u: User) -> Tuple[float, float, int, int]:
    usdt = _as_float(_get_user_field(u, "usdt_balance", "balance_usdt", default=0))
    ton  = _as_float(_get_user_field(u, "ton_balance", "balance_ton", default=0))
    pts  = int(_get_user_field(u, "point_balance", "balance_points", default=0) or 0)
    eng  = int(_get_user_field(u, "energy_balance", "energy", default=0) or 0)
    return usdt, ton, pts, eng


def _compute_last_seen_map(tg_ids: Optional[Sequence[int]] = None) -> Dict[int, datetime]:
    """
    计算每个用户的最后上线时间：
      last_seen_at = max(user.updated_at, MAX(ledger.created_at of user))
    可传 tg_ids 做范围过滤；传 None 表示全量。
    """
    last_seen: Dict[int, datetime] = {}

    # 取 user.updated_at
    with get_session() as s:
        uq = s.query(User.tg_id, User.updated_at)
        if tg_ids:
            uq = uq.filter(User.tg_id.in_(list(tg_ids)))
        for tgid, upd in uq.all():
            if tgid is None:
                continue
            if isinstance(upd, datetime):
                last_seen[int(tgid)] = upd

    # 取 ledger 最大时间
    with get_session() as s:
        lq = s.query(Ledger.user_tg_id, func.max(Ledger.created_at))
        if tg_ids:
            lq = lq.filter(Ledger.user_tg_id.in_(list(tg_ids)))
        lq = lq.group_by(Ledger.user_tg_id)
        for uid, mx in lq.all():
            if uid is None or not isinstance(mx, datetime):
                continue
            uid = int(uid)
            cur = last_seen.get(uid)
            last_seen[uid] = mx if (cur is None or mx > cur) else cur
    return last_seen


# ---------------------------------------
# 导出：所有用户“详细资料”总表（增强：新增 nickname/last_seen_at）
# ---------------------------------------

def export_all_users_detail(fmt: str = "xlsx") -> Optional[str]:
    with get_session() as s:
        users: List[User] = s.query(User).order_by(User.id.asc()).all()
    if not users:
        return None

    # 预先计算 last_seen
    tgt_ids = [int(getattr(u, "tg_id")) for u in users if getattr(u, "tg_id", None)]
    last_seen_map = _compute_last_seen_map(tgt_ids)

    recs: List[Dict[str, Any]] = []
    for u in users:
        tg_id = _get_user_field(u, "tg_id")
        username = _get_user_field(u, "username", "name", default="") or ""
        first_name = _get_user_field(u, "first_name", default="") or ""
        last_name  = _get_user_field(u, "last_name", default="") or ""
        full_name  = _get_user_field(u, "full_name", default="") or ""
        nickname   = _display_full_name(username, full_name, first_name, last_name, _get_user_field(u, "name", default=""))

        usdt, ton, pts, eng = _user_balances(u)
        last_seen = _dtfmt(last_seen_map.get(int(tg_id)) if tg_id else None)

        recs.append({
            "user_id": _get_user_field(u, "id"),
            "user_tg_id": tg_id,
            "username": username,
            "nickname": nickname,
            "language": _get_user_field(u, "language", "lang", default="") or "",
            "role": _enum_to_str(_get_user_field(u, "role", default="")),
            "usdt_balance": usdt,
            "ton_balance": ton,
            "point_balance": pts,
            "energy_balance": eng,
            "last_target_chat_id": _get_user_field(u, "last_target_chat_id", default=None),
            "last_target_chat_title": _get_user_field(u, "last_target_chat_title", default="") or "",
            "created_at": _dtfmt(_get_user_field(u, "created_at", default=None)),
            "updated_at": _dtfmt(_get_user_field(u, "updated_at", default=None)),
            "last_seen_at": last_seen,
        })

    cols = [
        "user_id","user_tg_id","username","nickname","language","role",
        "usdt_balance","ton_balance","point_balance","energy_balance",
        "last_target_chat_id","last_target_chat_title",
        "created_at","updated_at","last_seen_at",
    ]
    df = pd.DataFrame(recs, columns=cols)

    header_zh = [
        "用户自增ID","Telegram用户ID","用户名","用户昵称","语言","角色",
        "USDT余额","TON余额","积分余额","能量值",
        "最近目标群ID","最近目标群名称",
        "创建时间","更新时间","最后上线时间",
    ]
    path = _write_dataframe(
        df, basename="export_users_detail", fmt=fmt,
        header_zh=header_zh, sheet_name="records",
        freeze_header=True, enable_filter=True
    )
    return path


# ---------------------------------------
# 单用户专用 Excel（增强：Summary 的“最近活跃”采用 last_seen 逻辑；Ledger 增加 nickname）
# ---------------------------------------

def _collect_last7d_stats(uid_tg: int) -> Dict[str, Any]:
    since = datetime.utcnow() - timedelta(days=7)
    with get_session() as s:
        q = s.query(Ledger).filter(
            (Ledger.user_tg_id == uid_tg) if hasattr(Ledger, "user_tg_id") else (Ledger.user_id == uid_tg),
            Ledger.created_at >= since
        )
        rows = q.all()

    total = 0.0
    by_type: Dict[str, float] = {}
    for r in rows:
        amt = _as_float(getattr(r, "amount", 0))
        total += amt
        tname = _enum_to_str(getattr(r, "type", ""))
        by_type[tname] = by_type.get(tname, 0.0) + amt

    top_types = sorted(by_type.items(), key=lambda x: abs(x[1]), reverse=True)[:8]
    return {"days": 7, "total_amount": total, "top_types": top_types, "count": len(rows), "since": _dtfmt(since)}


def export_one_user_full(
    user_id_or_username: Union[int, str],
    fmt: str = "xlsx",
) -> Optional[str]:
    user = _pick_user_by_query(user_id_or_username)
    if not user:
        df = pd.DataFrame([{"error": f"user not found: {user_id_or_username}"}])
        return _write_dataframe(df, "export_user_not_found", fmt)

    # Ledger
    rows = _query_ledgers(user, start=None, end=None, tokens=None, types=None)
    recs = _rows_to_records(rows)
    _attach_user_columns(recs)
    ledger_cols = [
        "id", "created_at", "user_tg_id",
        "username", "first_name", "last_name", "full_name", "nickname",
        "type", "token", "amount", "note"
    ]
    df_ledger = pd.DataFrame(recs, columns=ledger_cols)
    ledger_header_zh = [
        "流水ID","创建时间","用户TGID",
        "用户名","名","姓","全名","用户昵称",
        "类型","代币","金额","备注"
    ]

    # Summary
    usdt, ton, pts, eng = _user_balances(user)
    last_chat_id = _get_user_field(user, "last_target_chat_id", default="")
    last_chat_title = _get_user_field(user, "last_target_chat_title", default="") or ""
    uname = _get_user_field(user, "username", "name", default="") or ""
    first_name = _get_user_field(user, "first_name", default="") or ""
    last_name  = _get_user_field(user, "last_name", default="") or ""
    full_name  = _get_user_field(user, "full_name", default="") or ""
    nickname   = _display_full_name(uname, full_name, first_name, last_name, _get_user_field(user, "name", default=""))

    # last_seen = max(updated_at, last ledger time)
    last_ledger_time = None
    if rows:
        try:
            last_ledger_time = max([getattr(r, "created_at") for r in rows if getattr(r, "created_at", None)])
        except Exception:
            last_ledger_time = None
    last_seen = max([dt for dt in [getattr(user, "updated_at", None), last_ledger_time] if isinstance(dt, datetime)], default=None)

    last7d = _collect_last7d_stats(int(getattr(user, "tg_id")))

    summary_rows = [
        {"字段": "用户自增ID", "值": _get_user_field(user, "id")},
        {"字段": "Telegram用户ID", "值": _get_user_field(user, "tg_id")},
        {"字段": "用户名", "值": uname},
        {"字段": "用户昵称", "值": nickname},
        {"字段": "语言", "值": _get_user_field(user, "language", "lang", default="")},
        {"字段": "角色", "值": _enum_to_str(_get_user_field(user, "role", default=""))},
        {"字段": "创建时间", "值": _dtfmt(_get_user_field(user, "created_at", default=None))},
        {"字段": "最近活跃", "值": _dtfmt(last_seen)},
        {"字段": "USDT余额", "值": f"{usdt:.2f}"},
        {"字段": "TON余额", "值": f"{ton:.2f}"},
        {"字段": "积分余额", "值": int(pts)},
        {"字段": "能量值", "值": int(eng)},
        {"字段": "最近目标群ID", "值": last_chat_id or ""},
        {"字段": "最近目标群名称", "值": last_chat_title or ""},
        {"字段": "近7天流水条数", "值": last7d["count"]},
        {"字段": "近7天金额合计", "值": f"{last7d['total_amount']:.4f}"},
        {"字段": "近7天统计起点", "值": last7d["since"]},
    ]
    for tname, amt in last7d["top_types"]:
        summary_rows.append({"字段": f"近7天-{tname}", "值": f"{amt:.4f}"})
    df_summary = pd.DataFrame(summary_rows, columns=["字段", "值"])

    # 写工作簿
    _ensure_export_dir()
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(EXPORT_DIR, f"user_{getattr(user, 'tg_id', 'unknown')}_{ts}.xlsx")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(list(df_summary.columns))
    for row in df_summary.itertuples(index=False, name=None):
        ws1.append(list(row))
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(df_summary.columns))}1"
    ws1.freeze_panes = "A2"
    _autosize_columns(ws1, start_row=1)

    ws2 = wb.create_sheet("Ledger")
    ws2.append(ledger_header_zh)
    ws2.append(ledger_cols)
    for row in df_ledger.itertuples(index=False, name=None):
        ws2.append(list(row))
    last_col_letter = get_column_letter(len(ledger_cols))
    ws2.auto_filter.ref = f"A2:{last_col_letter}2"
    ws2.freeze_panes = "A3"
    _autosize_columns(ws2, start_row=1)

    wb.save(path)
    return path


# ---------------------------------------
# 合并导出（全量）：Users + Ledger（增强：新增 nickname/last_seen_at）
# ---------------------------------------

def _build_users_dataframe(tg_ids: Optional[Sequence[int]] = None) -> pd.DataFrame:
    with get_session() as s:
        q = s.query(User)
        if tg_ids:
            q = q.filter(User.tg_id.in_(list(tg_ids)))
        users: List[User] = q.order_by(User.id.asc()).all()

    last_seen_map = _compute_last_seen_map([int(getattr(u, "tg_id")) for u in users if getattr(u, "tg_id", None)])

    recs: List[Dict[str, Any]] = []
    for u in users:
        tg_id = _get_user_field(u, "tg_id")
        username = _get_user_field(u, "username", "name", default="") or ""
        first_name = _get_user_field(u, "first_name", default="") or ""
        last_name  = _get_user_field(u, "last_name", default="") or ""
        full_name  = _get_user_field(u, "full_name", default="") or ""
        nickname   = _display_full_name(username, full_name, first_name, last_name, _get_user_field(u, "name", default=""))

        usdt, ton, pts, eng = _user_balances(u)
        last_seen = _dtfmt(last_seen_map.get(int(tg_id)) if tg_id else None)

        recs.append({
            "user_id": _get_user_field(u, "id"),
            "user_tg_id": tg_id,
            "username": username,
            "nickname": nickname,
            "language": _get_user_field(u, "language", "lang", default="") or "",
            "role": _enum_to_str(_get_user_field(u, "role", default="")),
            "usdt_balance": usdt,
            "ton_balance": ton,
            "point_balance": pts,
            "energy_balance": eng,
            "last_target_chat_id": _get_user_field(u, "last_target_chat_id", default=None),
            "last_target_chat_title": _get_user_field(u, "last_target_chat_title", default="") or "",
            "created_at": _dtfmt(_get_user_field(u, "created_at", default=None)),
            "updated_at": _dtfmt(_get_user_field(u, "updated_at", default=None)),
            "last_seen_at": last_seen,
        })

    cols = [
        "user_id","user_tg_id","username","nickname","language","role",
        "usdt_balance","ton_balance","point_balance","energy_balance",
        "last_target_chat_id","last_target_chat_title",
        "created_at","updated_at","last_seen_at",
    ]
    return pd.DataFrame(recs, columns=cols)


def _iter_all_ledgers_in_chunks(
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    tokens: Optional[Iterable[str]] = None,
    types: Optional[Iterable[Union[LedgerType, str]]] = None,
    chunk_size: int = 5000,
    tg_ids: Optional[Sequence[int]] = None,
) -> Generator[List[Dict[str, Any]], None, None]:
    """
    以分块方式遍历流水。可选 tg_ids 做范围过滤。
    每次产出一批“已附上用户字段（含 nickname）”的记录（dict 列表）。
    """
    offset = 0
    while True:
        with get_session() as s:
            q = s.query(Ledger)
            if tg_ids:
                q = q.filter(Ledger.user_tg_id.in_(list(tg_ids)))
            if start is not None:
                q = q.filter(Ledger.created_at >= start)
            if end is not None:
                q = q.filter(Ledger.created_at < end)
            if tokens:
                toks = [str(t).upper() for t in tokens]
                q = q.filter(Ledger.token.in_(toks))
            if types:
                norm_types: List[str] = []
                for t in types:
                    if isinstance(t, LedgerType):
                        norm_types.append(t.name)
                    else:
                        norm_types.append(str(t).upper())
                try:
                    q = q.filter(Ledger.type.in_(norm_types))
                except Exception:
                    back: List[LedgerType] = []
                    for n in norm_types:
                        try:
                            back.append(LedgerType[n])
                        except Exception:
                            pass
                    if back:
                        q = q.filter(Ledger.type.in_(back))
            q = q.order_by(Ledger.created_at.asc(), Ledger.id.asc()).offset(offset).limit(chunk_size)
            rows = q.all()
        if not rows:
            break
        recs = _rows_to_records(rows)
        _attach_user_columns(recs)
        yield recs
        if len(rows) < chunk_size:
            break
        offset += chunk_size


def export_all_users_and_ledger(
    fmt: str = "xlsx",
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    tokens: Optional[Iterable[str]] = None,
    types: Optional[Iterable[Union[LedgerType, str]]] = None,
    chunk_size: int = 5000,
) -> Optional[str]:
    """
    一个 Excel 两个工作表：
    - Sheet1: Users（所有用户详细信息；第1行中文表头，第2行英文列名；冻结+筛选）
    - Sheet2: Ledger（全部流水；第1行中文表头，第2行英文列名；冻结+筛选；分块写入）
    """
    df_users = _build_users_dataframe()
    if df_users is None or df_users.empty:
        df_users = pd.DataFrame(columns=[
            "user_id","user_tg_id","username","nickname","language","role",
            "usdt_balance","ton_balance","point_balance","energy_balance",
            "last_target_chat_id","last_target_chat_title",
            "created_at","updated_at","last_seen_at",
        ])

    ledger_cols = [
        "id", "created_at", "user_tg_id",
        "username", "first_name", "last_name", "full_name", "nickname",
        "type", "token", "amount", "note", "last_seen_at"
    ]
    users_header_zh = [
        "用户自增ID","Telegram用户ID","用户名","用户昵称","语言","角色",
        "USDT余额","TON余额","积分余额","能量值",
        "最近目标群ID","最近目标群名称",
        "创建时间","更新时间","最后上线时间",
    ]
    ledger_header_zh = [
        "流水ID","创建时间","用户TGID",
        "用户名","名","姓","全名","用户昵称",
        "类型","代币","金额","备注","最后上线时间"
    ]

    _ensure_export_dir()
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(EXPORT_DIR, f"export_all_{ts}.xlsx")

    # last_seen_map（全量）
    with get_session() as s:
        all_tgids = [r[0] for r in s.query(User.tg_id).filter(User.tg_id.isnot(None)).all()]
    last_seen_map = _compute_last_seen_map(all_tgids)

    wb = Workbook()

    # Users
    ws_users = wb.active
    ws_users.title = "Users"
    ws_users.append(users_header_zh)
    ws_users.append(list(df_users.columns))
    for row in df_users.itertuples(index=False, name=None):
        ws_users.append(list(row))
    ws_users.auto_filter.ref = f"A2:{get_column_letter(len(df_users.columns))}2"
    ws_users.freeze_panes = "A3"
    _autosize_columns(ws_users, start_row=1)

    # Ledger（分块）
    ws_ledger = wb.create_sheet("Ledger")
    ws_ledger.append(ledger_header_zh)
    ws_ledger.append(ledger_cols)
    for recs in _iter_all_ledgers_in_chunks(start=start, end=end, tokens=tokens, types=types, chunk_size=chunk_size):
        for r in recs:
            uid = r.get("user_tg_id")
            last_seen_str = _dtfmt(last_seen_map.get(int(uid)) if uid else None)
            row = [
                r.get("id"),
                r.get("created_at"),
                uid,
                r.get("username", ""),
                r.get("first_name", ""),
                r.get("last_name", ""),
                r.get("full_name", ""),
                r.get("nickname", ""),
                r.get("type", ""),
                r.get("token", ""),
                r.get("amount", 0),
                r.get("note", ""),
                last_seen_str,
            ]
            ws_ledger.append(row)
    ws_ledger.auto_filter.ref = f"A2:{get_column_letter(len(ledger_cols))}2"
    ws_ledger.freeze_panes = "A3"
    _autosize_columns(ws_ledger, start_row=1)

    wb.save(path)
    return path


# ---------------------------------------
# 新增：多用户合并导出（Users + Ledger，仅所选）
# ---------------------------------------

def export_users_full(
    tg_ids: List[int],
    fmt: str = "xlsx",
) -> Optional[str]:
    """
    一个 Excel 两个工作表（只包含所选用户）：
    - Sheet1: Users（所选用户）【含 nickname/last_seen_at】
    - Sheet2: Ledger（所选用户全部流水）【含 nickname/last_seen_at】
    """
    tg_ids = [int(x) for x in tg_ids if x is not None]
    if not tg_ids:
        return None

    df_users = _build_users_dataframe(tg_ids)
    if df_users is None or df_users.empty:
        return None

    # last_seen_map（仅所选）
    last_seen_map = _compute_last_seen_map(tg_ids)

    # Ledger 取数（所选用户）
    with get_session() as s:
        q = s.query(Ledger).filter(Ledger.user_tg_id.in_(tg_ids)).order_by(Ledger.created_at.asc(), Ledger.id.asc())
        rows: List[Ledger] = q.all()
    recs = _rows_to_records(rows)
    _attach_user_columns(recs)

    ledger_cols = [
        "id", "created_at", "user_tg_id",
        "username", "first_name", "last_name", "full_name", "nickname",
        "type", "token", "amount", "note", "last_seen_at"
    ]
    users_header_zh = [
        "用户自增ID","Telegram用户ID","用户名","用户昵称","语言","角色",
        "USDT余额","TON余额","积分余额","能量值",
        "最近目标群ID","最近目标群名称",
        "创建时间","更新时间","最后上线时间",
    ]
    ledger_header_zh = [
        "流水ID","创建时间","用户TGID",
        "用户名","名","姓","全名","用户昵称",
        "类型","代币","金额","备注","最后上线时间"
    ]

    _ensure_export_dir()
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(EXPORT_DIR, f"export_users_{ts}.xlsx")

    wb = Workbook()

    # Users
    ws_users = wb.active
    ws_users.title = "Users"
    ws_users.append(users_header_zh)
    ws_users.append(list(df_users.columns))
    for row in df_users.itertuples(index=False, name=None):
        ws_users.append(list(row))
    ws_users.auto_filter.ref = f"A2:{get_column_letter(len(df_users.columns))}2"
    ws_users.freeze_panes = "A3"
    _autosize_columns(ws_users, start_row=1)

    # Ledger
    ws_ledger = wb.create_sheet("Ledger")
    ws_ledger.append(ledger_header_zh)
    ws_ledger.append(ledger_cols)
    for r in recs:
        uid = r.get("user_tg_id")
        last_seen_str = _dtfmt(last_seen_map.get(int(uid)) if uid else None)
        row = [
            r.get("id"),
            r.get("created_at"),
            uid,
            r.get("username", ""),
            r.get("first_name", ""),
            r.get("last_name", ""),
            r.get("full_name", ""),
            r.get("nickname", ""),
            r.get("type", ""),
            r.get("token", ""),
            r.get("amount", 0),
            r.get("note", ""),
            last_seen_str,
        ]
        ws_ledger.append(row)
    ws_ledger.auto_filter.ref = f"A2:{get_column_letter(len(ledger_cols))}2"
    ws_ledger.freeze_panes = "A3"
    _autosize_columns(ws_ledger, start_row=1)

    wb.save(path)
    return path


# ---------------------------------------
# ★ 新增：admin 多选导出的对外主入口（名字务必与路由里调用的一致）
# ---------------------------------------

def export_some_users_and_ledger(*args, **kwargs) -> Optional[str]:
    """
    统一入口，供 admin 的 “multi export” 调用：

    兼容多种传参：
    - 位置参数：export_some_users_and_ledger([1,2,3]) 或 export_some_users_and_ledger("1,2,3")
    - 关键字参数：tg_ids=..., user_ids=..., ids=..., members=...（字符串或列表均可）
    - 可选 fmt（默认 xlsx）

    注意：此处认为传入的是 Telegram 用户 ID（tg_id）。如你的管理端传的是内部自增 ID，
    可在此处改为把内部 ID 映射到 tg_id 后再调用 export_users_full。
    """
    # 解析用户 ID 列表
    tg_ids: List[int] = []
    if args:
        tg_ids = _as_int_list(args[0])
    if not tg_ids:
        for key in ("tg_ids", "user_ids", "ids", "members"):
            if key in kwargs and kwargs.get(key) is not None:
                tg_ids = _as_int_list(kwargs.get(key))
                if tg_ids:
                    break
    fmt = str(kwargs.get("fmt", "xlsx")).lower()

    # 去重并排序，避免重复记录
    tg_ids = sorted(set(tg_ids))
    if not tg_ids:
        # 返回一份包含错误提示的 Excel，避免上层崩溃
        df = pd.DataFrame([{"error": "no user ids provided"}])
        return _write_dataframe(df, "export_empty_selection", fmt)

    # 复用已有能力
    return export_users_full(tg_ids=tg_ids, fmt=fmt)


# ---------------------------------------
# 便捷命令行测试（可选）
# ---------------------------------------
if __name__ == "__main__":
    import sys
    # 用法：python export_service.py 1,2,3  或  python export_service.py 1 2 3
    argv = sys.argv[1:]
    if len(argv) == 1 and ("," in argv[0] or ";" in argv[0]):
        ids = _as_int_list(argv[0])
    else:
        ids = _as_int_list(argv)
    path = export_some_users_and_ledger(ids)
    print(path if path else "<no file generated>")
